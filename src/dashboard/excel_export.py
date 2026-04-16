from __future__ import annotations

import re
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage


def safe_sheet_name(name: str) -> str:
    cleaned = re.sub(r'[:\\/?*\[\]]', "-", name)
    return cleaned[:31]


def write_excel(
    workbook_path: str | Path,
    df: pd.DataFrame,
    summaries: dict[str, pd.DataFrame],
    chart_paths: list[Path],
    insights_df: pd.DataFrame,
    recommendations_df: pd.DataFrame,
    trends_df: pd.DataFrame,
    popularity_df: pd.DataFrame,
    role_theme_df: pd.DataFrame,
) -> None:
    workbook_path = Path(workbook_path)

    with pd.ExcelWriter(workbook_path, engine="openpyxl") as writer:
        summaries["overview"].to_excel(writer, sheet_name="Overview", index=False)
        summaries["role_summary"].to_excel(writer, sheet_name="Role Summary", index=False)
        summaries["top_loved_specs"].to_excel(writer, sheet_name="Top Loved Specs", index=False)
        summaries["top_complained_specs"].to_excel(writer, sheet_name="Top Complained Specs", index=False)
        summaries["class_summary"].to_excel(writer, sheet_name="Class Summary", index=False)
        summaries["spec_summary"].to_excel(writer, sheet_name="Spec Summary", index=False)
        summaries["subreddit_summary"].to_excel(writer, sheet_name="Source Summary", index=False)
        summaries["theme_summary"].to_excel(writer, sheet_name="Theme Summary", index=False)
        summaries["class_theme_summary"].to_excel(writer, sheet_name="Class Themes", index=False)
        summaries["spec_theme_summary"].to_excel(writer, sheet_name="Spec Themes", index=False)
        summaries["positive_theme_summary"].to_excel(writer, sheet_name="Positive Themes", index=False)
        summaries["negative_theme_summary"].to_excel(writer, sheet_name="Negative Themes", index=False)

        insights_df.to_excel(writer, sheet_name="Design Insights", index=False)
        recommendations_df.to_excel(writer, sheet_name="Recommendations", index=False)
        trends_df.to_excel(writer, sheet_name="Time Trends", index=False)
        popularity_df.to_excel(writer, sheet_name="Popularity", index=False)
        role_theme_df.to_excel(writer, sheet_name="Role Theme Matrix", index=False)

        df.to_excel(writer, sheet_name="Raw Data", index=False)

        # Placeholder sheet for embedded charts
        pd.DataFrame({"Charts": ["See embedded visuals below"]}).to_excel(writer, sheet_name="Charts", index=False)

        for wow_class in sorted(df["wow_class"].dropna().unique()):
            cdf = df[df["wow_class"] == wow_class]
            cdf.to_excel(writer, sheet_name=safe_sheet_name(f"{wow_class} Data"), index=False)

        for (wow_class, spec), sdf in df.groupby(["wow_class", "spec"]):
            sdf.to_excel(writer, sheet_name=safe_sheet_name(f"{spec} {wow_class}"), index=False)

    wb = load_workbook(workbook_path)

    # Overview quick tables
    ws = wb["Overview"]

    ws["J4"] = "Top Loved Specs"
    ws["J15"] = "Top Complained Specs"
    ws["J26"] = "Role Summary"
    ws["J36"] = "Top Positive Themes"
    ws["J47"] = "Top Negative Themes"
    ws["J58"] = "Top Recommendations"

    loved = summaries.get("top_loved_specs", pd.DataFrame())
    complained = summaries.get("top_complained_specs", pd.DataFrame())
    roles = summaries.get("role_summary", pd.DataFrame())
    pos_themes = summaries.get("positive_theme_summary", pd.DataFrame())
    neg_themes = summaries.get("negative_theme_summary", pd.DataFrame())

    start_row = 5
    if not loved.empty:
        ws[f"J{start_row}"] = "Class"
        ws[f"K{start_row}"] = "Spec"
        ws[f"L{start_row}"] = "Avg Sentiment"
        ws[f"M{start_row}"] = "Records"
        for i, (_, row) in enumerate(loved.head(5).iterrows(), start=1):
            ws[f"J{start_row + i}"] = row["wow_class"]
            ws[f"K{start_row + i}"] = row["spec"]
            ws[f"L{start_row + i}"] = float(row["avg_sentiment"])
            ws[f"M{start_row + i}"] = int(row["records"])

    start_row = 16
    if not complained.empty:
        ws[f"J{start_row}"] = "Class"
        ws[f"K{start_row}"] = "Spec"
        ws[f"L{start_row}"] = "Avg Sentiment"
        ws[f"M{start_row}"] = "Records"
        for i, (_, row) in enumerate(complained.head(5).iterrows(), start=1):
            ws[f"J{start_row + i}"] = row["wow_class"]
            ws[f"K{start_row + i}"] = row["spec"]
            ws[f"L{start_row + i}"] = float(row["avg_sentiment"])
            ws[f"M{start_row + i}"] = int(row["records"])

    start_row = 27
    if not roles.empty:
        ws[f"J{start_row}"] = "Role"
        ws[f"K{start_row}"] = "Avg Sentiment"
        ws[f"L{start_row}"] = "Records"
        for i, (_, row) in enumerate(roles.iterrows(), start=1):
            ws[f"J{start_row + i}"] = row["role"]
            ws[f"K{start_row + i}"] = float(row["avg_sentiment"])
            ws[f"L{start_row + i}"] = int(row["records"])

    start_row = 37
    if not pos_themes.empty:
        ws[f"J{start_row}"] = "Theme"
        ws[f"K{start_row}"] = "Count"
        for i, (_, row) in enumerate(pos_themes.head(8).iterrows(), start=1):
            ws[f"J{start_row + i}"] = row["theme"]
            ws[f"K{start_row + i}"] = int(row["count"])

    start_row = 48
    if not neg_themes.empty:
        ws[f"J{start_row}"] = "Theme"
        ws[f"K{start_row}"] = "Count"
        for i, (_, row) in enumerate(neg_themes.head(8).iterrows(), start=1):
            ws[f"J{start_row + i}"] = row["theme"]
            ws[f"K{start_row + i}"] = int(row["count"])

    start_row = 59
    if not recommendations_df.empty:
        ws[f"J{start_row}"] = "Class"
        ws[f"K{start_row}"] = "Spec"
        ws[f"L{start_row}"] = "Priority"
        ws[f"M{start_row}"] = "Recommendation"
        for i, (_, row) in enumerate(recommendations_df.head(5).iterrows(), start=1):
            ws[f"J{start_row + i}"] = row["wow_class"]
            ws[f"K{start_row + i}"] = row["spec"]
            ws[f"L{start_row + i}"] = row["priority"]
            ws[f"M{start_row + i}"] = row["recommendation"]

    # Dedicated charts sheet
    chart_ws = wb["Charts"]
    chart_ws["A1"] = "Dashboard Visuals"

    row_positions = [3, 25, 47, 69, 91, 113, 135, 157]
    for idx, image_path in enumerate(chart_paths[:8]):
        if Path(image_path).exists() and idx < len(row_positions):
            img = XLImage(str(image_path))
            img.width = 900
            img.height = 420
            chart_ws.add_image(img, f"A{row_positions[idx]}")

    wb.save(workbook_path)